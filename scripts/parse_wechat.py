#!/usr/bin/env python3
"""
Parse WeChat bill CSV/XLSX to intermediate JSON format.
AI will use this JSON for further processing (category prediction, etc).

Output JSON format: same as parse_alipay.py
"""

import csv
import json
import sys
from datetime import datetime
from pathlib import Path

SKILL_DIR = Path(__file__).parent.parent
CONFIG_DIR = SKILL_DIR / "config"


def load_config():
    """Load all config files."""
    config = {}
    for name in ["account_map", "category_rules", "payee_rules", "description_rules"]:
        config_file = CONFIG_DIR / f"{name}.json"
        if config_file.exists():
            with open(config_file, "r", encoding="utf-8") as f:
                config[name] = json.load(f)
        else:
            config[name] = {}
    return config


def map_account(payment_method: str, config: dict) -> str:
    """Map payment method to MoneyWiz account name."""
    if not payment_method or payment_method.strip() in ["", "/"]:
        return config.get("account_map", {}).get("defaults", {}).get("wechat", "微信零钱")

    account_map = config.get("account_map", {})

    for card_suffix, account_name in account_map.get("bank_cards", {}).items():
        if card_suffix in payment_method:
            return account_name

    for keyword, account_name in account_map.get("keywords", {}).items():
        if keyword in payment_method:
            return account_name

    return config.get("account_map", {}).get("defaults", {}).get("wechat", "微信零钱")


def predict_category(payee: str, description: str, original_category: str, config: dict) -> tuple:
    """Predict category. Returns (category, confidence).

    Priority: description keywords > payee > original category
    This ensures specific items like coffee are correctly categorized
    even when ordered through platforms like Meituan.
    """
    rules = config.get("category_rules", {})

    # 1. Keyword match in description (highest priority)
    for keyword, category in rules.get("keyword_category", {}).items():
        if keyword in description:
            return category, "high"

    # 2. Exact match payee
    for payee_key, category in rules.get("payee_category", {}).items():
        if payee_key in payee:
            return category, "high"

    # 3. Keyword match in payee (fallback)
    for keyword, category in rules.get("keyword_category", {}).items():
        if keyword in payee:
            return category, "medium"

    # 4. Use original if available
    if original_category and original_category.strip() and original_category.strip() != "/":
        return original_category, "low"

    return "", "none"


def simplify_payee(payee: str, config: dict) -> str:
    """Simplify payee name."""
    if not payee:
        return ""

    rules = config.get("payee_rules", {})

    for original, simplified in rules.get("simplify", {}).items():
        if original in payee:
            return simplified

    result = payee
    for suffix in ["（上海）", "(上海)", "有限公司", "有限责任公司", "科技", "网络", "信息", "服务"]:
        result = result.replace(suffix, "")

    return result.strip() or payee


def apply_payee_description(payee: str, description: str, config: dict) -> str:
    """Override description based on payee name if rule exists."""
    if not payee:
        return description

    rules = config.get("description_rules", {})
    payee_desc = rules.get("payee_description", {})

    for payee_key, new_desc in payee_desc.items():
        if payee_key.startswith("_"):
            continue
        if payee_key in payee:
            return new_desc

    return description


def simplify_description(description: str, config: dict) -> str:
    """Simplify transaction description by removing noise."""
    import re

    if not description:
        return ""

    rules = config.get("description_rules", {})
    result = description

    # Check if should keep prefix (e.g. "退款-", "转账备注:")
    keep_prefixes = rules.get("keep_prefixes", [])
    prefix_to_restore = ""
    for prefix in keep_prefixes:
        if result.startswith(prefix):
            prefix_to_restore = prefix
            result = result[len(prefix):]
            break

    # Apply regex removal patterns
    for pattern in rules.get("remove_patterns", []):
        try:
            result = re.sub(pattern, "", result)
        except re.error:
            pass

    # Apply replacements
    for old, new in rules.get("replacements", {}).items():
        try:
            result = re.sub(old, new, result)
        except re.error:
            if old in result:
                result = result.replace(old, new)

    # Remove suffixes like "等多件"
    for suffix_pattern in rules.get("simplify_suffixes", []):
        try:
            result = re.sub(suffix_pattern, "", result)
        except re.error:
            pass

    # Restore prefix
    if prefix_to_restore:
        result = prefix_to_restore + result

    return result.strip() or description


def format_date(date_str: str) -> str:
    """Format date to YYYY/MM/DD HH:mm:ss."""
    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M"]:
        try:
            dt = datetime.strptime(str(date_str).strip(), fmt)
            return dt.strftime("%Y/%m/%d %H:%M:%S")
        except ValueError:
            continue
    return str(date_str)


def parse_amount(amount_str: str) -> float:
    """Parse amount string, handling ¥ prefix."""
    if not amount_str:
        return 0.0
    cleaned = str(amount_str).strip().lstrip("¥").strip()
    try:
        return float(cleaned.replace(",", ""))
    except ValueError:
        return 0.0


def parse_wechat_row(row: dict, config: dict) -> dict:
    """Parse a single WeChat transaction row."""
    date = format_date(row.get("交易时间", ""))
    product = str(row.get("商品", "")).strip()
    tx_type = str(row.get("交易类型", "")).strip()
    payee_original = str(row.get("交易对方", "")).strip()
    income_expense = str(row.get("收/支", "")).strip()
    amount_str = str(row.get("金额(元)", "0")).strip()
    payment_method = str(row.get("支付方式", "")).strip()

    description_original = product if product and product != "/" else tx_type
    description = simplify_description(description_original, config)
    amount = parse_amount(amount_str)
    payee = simplify_payee(payee_original, config)
    # Apply payee-based description override
    description = apply_payee_description(payee, description, config)
    account = map_account(payment_method, config)

    tx = {
        "date": date,
        "description_original": description_original,
        "description": description,
        "payee_original": payee_original,
        "payee": payee,
        "account": account,
        "transfer_account": "",
        "original_category": tx_type,
        "payment_method": payment_method,
    }

    # Detect refund: tx_type ends with "-退款" or equals "退款"
    is_refund = "退款" in tx_type and (tx_type.endswith("退款") or "-退款" in tx_type)

    if is_refund:
        # Refund: imported as Income (Z_ENT=37), later converted to Refund (Z_ENT=43)
        # by convert_refunds.py. Description format: "<simplified> 的退款"
        tx["description"] = f"{description} 的退款"
        category, confidence = predict_category(payee_original, description, "", config)
        tx["category"] = category
        tx["category_confidence"] = confidence
        tx["amount"] = str(abs(amount))
        tx["is_refund"] = True
    elif income_expense == "其他":
        tx["transfer_account"] = map_account(payee_original, config)
        tx["category"] = ""
        tx["category_confidence"] = "high"
        tx["amount"] = str(-abs(amount)) if "还款" in description else str(amount)
    else:
        category, confidence = predict_category(payee_original, description, tx_type, config)
        tx["category"] = category
        tx["category_confidence"] = confidence
        tx["amount"] = str(-abs(amount)) if income_expense == "支出" else str(amount)

    tx["needs_review"] = tx["category_confidence"] in ["low", "none"]
    return tx


def parse_wechat_csv(file_path: str, config: dict) -> list:
    """Parse WeChat CSV file."""
    with open(file_path, "r", encoding="utf-8") as f:
        content = f.read()

    lines = content.split("\n")
    data_lines = []
    found_header = False

    for line in lines:
        if line.startswith("--"):
            continue
        if not found_header and "交易时间" in line:
            found_header = True
        if found_header and line.strip():
            data_lines.append(line)

    if not data_lines:
        raise ValueError("Cannot find data section in WeChat CSV")

    transactions = []
    reader = csv.DictReader("\n".join(data_lines).split("\n"))
    for row in reader:
        transactions.append(parse_wechat_row(row, config))

    return transactions


def parse_wechat_xlsx(file_path: str, config: dict) -> list:
    """Parse WeChat XLSX file."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl required: pip install openpyxl")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    header_row = None
    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and "交易时间" in row:
            header_row = i
            headers = list(row)
            break

    if not header_row:
        raise ValueError("Cannot find header row in WeChat XLSX")

    transactions = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not any(row):
            continue

        row_dict = {}
        for i, header in enumerate(headers):
            if header and i < len(row):
                row_dict[header] = row[i] if row[i] is not None else ""

        if row_dict.get("交易时间"):
            transactions.append(parse_wechat_row(row_dict, config))

    return transactions


def parse_wechat(file_path: str, config: dict) -> dict:
    """Parse WeChat bill and return structured data."""
    file_path = Path(file_path)
    ext = file_path.suffix.lower()

    if ext == ".xlsx":
        transactions = parse_wechat_xlsx(str(file_path), config)
    elif ext == ".csv":
        transactions = parse_wechat_csv(str(file_path), config)
    else:
        raise ValueError(f"Unsupported format: {ext}")

    high = sum(1 for tx in transactions if tx["category_confidence"] == "high")
    medium = sum(1 for tx in transactions if tx["category_confidence"] == "medium")
    needs_review = sum(1 for tx in transactions if tx["needs_review"])

    return {
        "source": "wechat",
        "file": file_path.name,
        "parsed_at": datetime.now().isoformat(),
        "transactions": transactions,
        "stats": {
            "total": len(transactions),
            "high_confidence": high,
            "medium_confidence": medium,
            "needs_review": needs_review
        }
    }


def main():
    if len(sys.argv) < 2:
        print("Usage: python parse_wechat.py <input_file> [output_json]")
        print("\nParse WeChat bill (CSV/XLSX) to JSON for AI processing.")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) >= 3 else None

    config = load_config()
    result = parse_wechat(input_file, config)

    if output_file:
        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f"Output: {output_file}")
    else:
        print(json.dumps(result, ensure_ascii=False, indent=2))

    stats = result["stats"]
    print(f"\n[Parsed {stats['total']} transactions, {stats['needs_review']} need review]", file=sys.stderr)


if __name__ == "__main__":
    main()
